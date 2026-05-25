"""
Exporters: Convert structured data to JSON, CSV, Excel formats.
"""

import json
import csv
import io
from typing import Dict, Any, List
from datetime import datetime

# Pre-import for Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class JSONExporter:
    """Export structured data as JSON."""

    @staticmethod
    def export(data: Dict[str, Any], pretty: bool = True) -> str:
        """
        Export to JSON string.
        
        Args:
            data: Extracted structured data
            pretty: Pretty-print with indentation
            
        Returns:
            JSON string
        """
        if pretty:
            return json.dumps(data, indent=2, ensure_ascii=False, default=str)
        return json.dumps(data, ensure_ascii=False, default=str)

    @staticmethod
    def save_to_file(data: Dict[str, Any], filepath: str) -> str:
        """Save to JSON file."""
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        return f"JSON exported to {filepath}"


class CSVExporter:
    """Export structured data as CSV."""

    @staticmethod
    def export_records(records: List[Dict[str, Any]]) -> str:
        """
        Convert records to CSV.
        
        Args:
            records: List of record dicts (e.g., person records)
            
        Returns:
            CSV string
        """
        if not records:
            return ""
        
        output = io.StringIO()
        
        # Get all unique keys
        all_keys = set()
        for record in records:
            all_keys.update(record.keys())
        
        fieldnames = sorted(list(all_keys))
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        
        writer.writeheader()
        for record in records:
            writer.writerow(record)
        
        return output.getvalue()

    @staticmethod
    def export_tables(tables: List[Dict[str, Any]]) -> str:
        """
        Convert tables to CSV format.
        Returns one CSV block per table (separated by blank lines).
        """
        output = io.StringIO()
        
        for table_idx, table in enumerate(tables):
            headers = table.get("headers", [])
            rows = table.get("rows", [])
            
            if headers:
                output.write(",".join(headers) + "\n")
            
            for row in rows:
                output.write(",".join(str(cell) for cell in row) + "\n")
            
            # Separator between tables
            if table_idx < len(tables) - 1:
                output.write("\n")
        
        return output.getvalue()

    @staticmethod
    def save_to_file(csv_content: str, filepath: str) -> str:
        """Save CSV to file."""
        with open(filepath, "w", encoding="utf-8", newline="") as f:
            f.write(csv_content)
        return f"CSV exported to {filepath}"


class ExcelExporter:
    """Export structured data to Excel workbooks."""

    @staticmethod
    def export(extracted_data: Dict[str, Any], filename: str = "export.xlsx") -> bytes:
        """
        Create Excel workbook with multiple sheets.
        
        Args:
            extracted_data: All extracted structured data
            filename: Output filename
            
        Returns:
            Excel file bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 0: Summary (Overview of what was extracted)
        ws = wb.create_sheet("Summary", 0)
        ExcelExporter._write_summary_sheet(ws, extracted_data)
        
        # Sheet 1: Person Records (Primary - shows up first)
        records = extracted_data.get("records", [])
        if records:
            ws = wb.create_sheet("Records")
            ExcelExporter._write_records_sheet(ws, records)
        
        # Sheet 2: Metadata
        metadata = extracted_data.get("document_metadata", {})
        if metadata:
            ws = wb.create_sheet("Metadata")
            ExcelExporter._write_metadata_sheet(ws, metadata)
        
        # Sheet 3: Tables
        tables = extracted_data.get("tables", [])
        for idx, table in enumerate(tables, 1):
            ws = wb.create_sheet(f"Table_{idx}")
            ExcelExporter._write_table_sheet(ws, table)
        
        # Sheet 4: Key-Value Pairs
        kvp = extracted_data.get("key_value_pairs", {})
        if kvp:
            ws = wb.create_sheet("Form_Fields")
            ExcelExporter._write_kvp_sheet(ws, kvp)
        
        # Sheet 5: Lists
        lists = extracted_data.get("lists", [])
        for idx, lst in enumerate(lists, 1):
            ws = wb.create_sheet(f"List_{idx}")
            ExcelExporter._write_list_sheet(ws, lst)
        
        # Sheet 6: Items
        items = extracted_data.get("items", [])
        if items:
            ws = wb.create_sheet("Items")
            ExcelExporter._write_items_sheet(ws, items)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _write_summary_sheet(ws, extracted_data: Dict[str, Any]):
        """Write a summary sheet showing what was extracted."""
        ws["A1"] = "Extraction Summary"
        ws["A1"].font = Font(bold=True, size=14)
        
        row = 3
        
        # Count records
        records = extracted_data.get("records", [])
        ws[f"A{row}"] = "Person Records Found:"
        ws[f"B{row}"] = len(records)
        row += 1
        
        # Count tables
        tables = extracted_data.get("tables", [])
        ws[f"A{row}"] = "Tables Found:"
        ws[f"B{row}"] = len(tables)
        row += 1
        
        # Count key-value pairs
        kvp = extracted_data.get("key_value_pairs", {})
        ws[f"A{row}"] = "Form Fields Found:"
        ws[f"B{row}"] = len(kvp)
        row += 1
        
        # Count lists
        lists = extracted_data.get("lists", [])
        ws[f"A{row}"] = "Lists Found:"
        ws[f"B{row}"] = len(lists)
        row += 2
        
        # Show metadata
        metadata = extracted_data.get("document_metadata", {})
        if metadata:
            ws[f"A{row}"] = "Document Metadata"
            ws[f"A{row}"].font = Font(bold=True, size=11)
            row += 1
            
            for key, value in metadata.items():
                ws[f"A{row}"] = key.replace("_", " ").title()
                ws[f"B{row}"] = str(value)
                row += 1
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50

    @staticmethod
    def _write_metadata_sheet(ws, metadata: Dict[str, Any]):
        """Write metadata to worksheet."""
        ws["A1"] = "Property"
        ws["B1"] = "Value"
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        
        row = 2
        for key, value in metadata.items():
            ws[f"A{row}"] = key
            ws[f"B{row}"] = value
            row += 1
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    @staticmethod
    def _write_records_sheet(ws, records: List[Dict[str, Any]]):
        """Write person records to worksheet with proper formatting."""
        if not records:
            return
        
        # Get all unique keys from all records to ensure we capture all columns
        all_keys = set()
        for record in records:
            all_keys.update(record.keys())
        
        # Sort keys for consistent column order (predictable and nice)
        # Priority order: no, id, name, age, gender, phone, email, class, department, then others
        priority_order = ["no", "id", "name", "age", "gender", "phone", "email", "contact_number", "class", "department"]
        headers = []
        for key in priority_order:
            if key in all_keys:
                headers.append(key)
                all_keys.discard(key)
        # Add remaining keys in sorted order
        headers.extend(sorted(all_keys))
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data rows
        for row_idx, record in enumerate(records, 2):
            for col, header in enumerate(headers, 1):
                value = record.get(header, "")
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Auto-width columns (with minimum widths for readability)
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            # Calculate width based on header length and content
            max_length = len(headers[col - 1])
            for record in records:
                cell_value = str(record.get(headers[col - 1], ""))
                max_length = max(max_length, len(cell_value))
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 40)  # Cap at 40 for very long content
            adjusted_width = max(adjusted_width, 12)  # Minimum width
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"

    @staticmethod
    def _write_table_sheet(ws, table: Dict[str, Any]):
        """Write table data to worksheet with proper formatting."""
        headers = table.get("headers", [])
        rows = table.get("rows", [])
        
        if not headers:
            return
        
        # Headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Rows with alternating colors for readability
        light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for row_idx, row_data in enumerate(rows, 2):
            fill = light_fill if row_idx % 2 == 0 else white_fill
            for col, cell_data in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=cell_data)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Auto-width columns
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            max_length = len(str(headers[col - 1]))
            for row in rows:
                if col <= len(row):
                    max_length = max(max_length, len(str(row[col - 1])))
            
            adjusted_width = min(max_length + 2, 40)
            adjusted_width = max(adjusted_width, 12)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"

    @staticmethod
    def _write_kvp_sheet(ws, kvp: Dict[str, str]):
        """Write key-value pairs to worksheet."""
        ws["A1"] = "Key"
        ws["B1"] = "Value"
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        
        row = 2
        for key, value in kvp.items():
            ws[f"A{row}"] = key
            ws[f"B{row}"] = value
            row += 1
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50

    @staticmethod
    def _write_list_sheet(ws, lst: Dict[str, Any]):
        """Write list items to worksheet."""
        list_type = lst.get("type", "list")
        items = lst.get("items", [])
        
        ws["A1"] = f"{list_type.replace('_', ' ').title()}"
        ws["A1"].font = Font(bold=True, size=12)
        
        row = 2
        for idx, item in enumerate(items, 1):
            ws[f"A{row}"] = f"{idx}. {item}"
            row += 1
        
        ws.column_dimensions["A"].width = 60

    @staticmethod
    def _write_items_sheet(ws, items: List[Dict[str, Any]]):
        """Write items (invoice/receipt) to worksheet."""
        headers = ["Description", "Quantity", "Price", "Subtotal"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        row = 2
        total = 0
        for item in items:
            ws.cell(row=row, column=1, value=item.get("description", ""))
            ws.cell(row=row, column=2, value=item.get("quantity", 0))
            ws.cell(row=row, column=3, value=item.get("price", 0))
            ws.cell(row=row, column=4, value=item.get("subtotal", 0))
            total += item.get("subtotal", 0)
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=4, value=total)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4).font = Font(bold=True)
        
        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12

    @staticmethod
    def save_to_file(excel_bytes: bytes, filepath: str) -> str:
        """Save Excel file."""
        with open(filepath, "wb") as f:
            f.write(excel_bytes)
        return f"Excel exported to {filepath}"
